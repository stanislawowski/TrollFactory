#!/usr/bin/env python3
# -*- coding: utf8 -*-
from fpdf import FPDF
from csv import writer
from openpyxl import Workbook

def fix_title(s):
    return s.replace('_', ' ').capitalize()

def generate_pdf(personality, id):
    pdf = FPDF(format='A4')
    pdf.add_page()
    pdf.add_font('PTSans', '', 'static/fonts/PTSans-Regular.ttf', uni=True)
    pdf.add_font('PTSans', 'B', 'static/fonts/PTSans-Bold.ttf', uni=True)

    pdf.image('static/img/pdf_header.png', 0, 0, 210)
    pdf.set_font('PTSans', 'B', 12)
    pdf.cell(60, 25, 'https://trollfactory.tk/' + str(id), 0, 1, '')
    
    for i in range(10):
        pdf.image('static/img/dna_strand.png', 180, 0+i*31.8, 13)

    for prop_name in personality.keys():
        pdf.set_font('PTSans', 'B', 12)
        pdf.cell(60, 6, personality[prop_name]['prop_title'], 0, 1, 'C')

        pdf.set_font('PTSans', '', 12)
        for prop_entry in personality[prop_name].keys():
            if prop_entry == 'user_agent':
                pdf.cell(0, 6, 'User-agent:', 0, 1)
                for part in personality[prop_name][prop_entry]:
                    pdf.cell(0, 6, str(' '*20+part), 0, 1)
            elif prop_entry != 'prop_title':
                line = fix_title(prop_entry)+': '+str(personality[prop_name][prop_entry])
                pdf.cell(0, 6, line, 0, 1)

        pdf.ln(4)
    
    pdf.output(''.join(['personalities/', str(id), '.pdf']), 'F')

def generate_csv(personality, id):
    path = ''.join(['personalities/', str(id), '.csv'])
    with open(path, 'w', encoding='utf-8') as file:
        wtr = writer(file)
        out_keys = []
        out_data = []

        for prop_name in personality.keys():
            for prop_entry in personality[prop_name].keys():
                if prop_entry != 'prop_title':
                    if prop_entry == 'user_agent':
                        personality[prop_name][prop_entry] = ' '.join(personality[prop_name][prop_entry])
                    out_keys.append(prop_name + '__' + prop_entry)
                    out_data.append(str(personality[prop_name][prop_entry]))

        wtr.writerow(out_keys)
        wtr.writerow(out_data)

def generate_xlsx(personality, id):
    xlsx = Workbook()
    xlsx.remove(xlsx.active)

    for prop_name in personality.keys():
        data = [[], []]
        sheet = xlsx.create_sheet(title = personality[prop_name]['prop_title'])

        for prop_entry in personality[prop_name].keys():
            if prop_entry == 'user_agent':
                personality[prop_name][prop_entry] = ' '.join(personality[prop_name][prop_entry])
            data[0].append(fix_title(prop_entry))
            data[1].append(str(personality[prop_name][prop_entry]))

        for row in data:
            sheet.append(row)

    xlsx.save(''.join(['personalities/', str(id), '.xlsx']))
